"""Generation des fichiers Excel et PDF : liste d'etudiants, bulletins et resultats."""
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from scolarite.models import Etablissement

ENTETE_COULEUR_XLSX = "1C2B45"       # openpyxl : hex sans #
ENTETE_COULEUR_PDF = "#1C2B45"       # reportlab : hex avec #


def _enseignants_du_bulletin(evaluations):
    """Liste (triee, sans doublon) des enseignants concernes par les evaluations
    d'un bulletin, via evaluation.objet.matiere -> Enseignant."""
    from .calculs import _enseignants_de_la_matiere
    noms = set()
    for ev in evaluations:
        matiere = getattr(ev['objet'], 'matiere', None)
        for nom in _enseignants_de_la_matiere(matiere):
            noms.add(nom)
    return sorted(noms)


# ============================================================
# Liste d'etudiants (sans notes)
# ============================================================

def _lignes_etudiants(queryset):
    lignes = []
    for u in queryset.select_related('profil'):
        profil = getattr(u, 'profil', None)
        lignes.append([
            u.last_name or '-',
            u.first_name or '-',
            u.username,
            profil.get_role_display() if profil else '-',
            str(profil.groupe) if profil and profil.groupe else '-',
            str(profil.classe) if profil and profil.classe else '-',
            str(profil.filiere) if profil and profil.filiere else '-',
            u.email or '-',
            profil.numero_telephone if profil else '-',
        ])
    return lignes


def exporter_etudiants_excel(queryset):
    entetes = ['Nom', 'Prenom', 'Identifiant', 'Role', 'Groupe', 'Classe', 'Filiere', 'Email', 'Telephone']
    lignes = _lignes_etudiants(queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etudiants"
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ENTETE_COULEUR_XLSX)
    for ligne in lignes:
        ws.append(ligne)
    for col in ws.columns:
        largeur = max(len(str(c.value)) for c in col if c.value) + 2
        ws.column_dimensions[col[0].column_letter].width = min(largeur, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etudiants.xlsx"'
    wb.save(response)
    return response


def exporter_etudiants_pdf(queryset):
    entetes = ['Nom', 'Prenom', 'Identifiant', 'Role', 'Groupe', 'Classe', 'Filiere', 'Email']
    lignes = [l[:8] for l in _lignes_etudiants(queryset)]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etudiants.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Liste des etudiants", styles['Title']), Spacer(1, 12)]

    table = Table([entetes] + lignes, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(ENTETE_COULEUR_PDF)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f2ea")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ============================================================
# Bulletin (notes + coefficients + moyenne + rang)
# ============================================================

def exporter_bulletin_excel(periode, groupe, evaluations, lignes):
    entetes = ['Rang', 'Nom', 'Prenom', 'Groupe']
    for ev in evaluations:
        entetes.append(f"{ev['label']} (coef {ev['coefficient']})")
    entetes += ['Moyenne /20', 'Moyenne /10']

    etablissement = Etablissement.charger()
    enseignants = _enseignants_du_bulletin(evaluations)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulletin"

    ligne_courante = 1
    if etablissement.nom:
        ws.append([etablissement.nom + (f" - Tél : {etablissement.telephone}" if etablissement.telephone else "")])
        ws.merge_cells(start_row=ligne_courante, start_column=1, end_row=ligne_courante, end_column=len(entetes))
        ws.cell(row=ligne_courante, column=1).font = Font(bold=True, size=12)
        ligne_courante += 1

    titre = f"Bulletin - {periode.nom}" + (f" ({periode.annee_scolaire})" if periode.annee_scolaire else "") + (f" - {groupe.nom}" if groupe else "")
    ws.append([titre])
    ws.merge_cells(start_row=ligne_courante, start_column=1, end_row=ligne_courante, end_column=len(entetes))
    ws.cell(row=ligne_courante, column=1).font = Font(bold=True, size=14)
    ligne_courante += 1

    if enseignants:
        ws.append(["Enseignant(s) : " + ", ".join(enseignants)])
        ws.merge_cells(start_row=ligne_courante, start_column=1, end_row=ligne_courante, end_column=len(entetes))
        ligne_courante += 1

    ws.append([])
    ligne_courante += 1
    ligne_entetes = ligne_courante + 1
    ws.append(entetes)
    for cell in ws[ligne_entetes]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ENTETE_COULEUR_XLSX)
        cell.alignment = Alignment(wrap_text=True)

    for ligne in lignes:
        profil = getattr(ligne['etudiant'], 'profil', None)
        row = [
            ligne['rang'] if ligne['rang'] is not None else '-',
            ligne['etudiant'].last_name or '-',
            ligne['etudiant'].first_name or '-',
            str(profil.groupe) if profil and profil.groupe else '-',
        ]
        for note in ligne['notes']:
            row.append(round(float(note), 2) if note is not None else '-')
        row.append(float(ligne['moyenne_20']) if ligne['moyenne_20'] is not None else '-')
        row.append(float(ligne['moyenne_10']) if ligne['moyenne_10'] is not None else '-')
        ws.append(row)

    from openpyxl.utils import get_column_letter
    for idx in range(1, len(entetes) + 1):
        lettre = get_column_letter(idx)
        valeurs = [str(ws.cell(row=r, column=idx).value) for r in range(ligne_entetes, ws.max_row + 1) if ws.cell(row=r, column=idx).value not in (None, '')]
        if valeurs:
            ws.column_dimensions[lettre].width = min(max(len(v) for v in valeurs) + 2, 30)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bulletin.xlsx"'
    wb.save(response)
    return response


def exporter_bulletin_pdf(periode, groupe, evaluations, lignes):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bulletin.pdf"'

    etablissement = Etablissement.charger()
    enseignants = _enseignants_du_bulletin(evaluations)

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet()
    elements = []

    if etablissement.nom:
        entete_etab = etablissement.nom + (f" — Tél : {etablissement.telephone}" if etablissement.telephone else "")
        elements.append(Paragraph(entete_etab, styles['Heading3']))

    titre = f"Bulletin - {periode.nom}" + (f" ({periode.annee_scolaire})" if periode.annee_scolaire else "") + (f" - {groupe.nom}" if groupe else " - Tous groupes")
    elements.append(Paragraph(titre, styles['Title']))

    if enseignants:
        elements.append(Paragraph("Enseignant(s) : " + ", ".join(enseignants), styles['Normal']))

    elements.append(Spacer(1, 10))

    entetes = ['Rang', 'Nom', 'Prenom'] + [f"{ev['label']}\n(coef {ev['coefficient']})" for ev in evaluations] + ['Moy. /20', 'Moy. /10']
    data = [entetes]
    for ligne in lignes:
        row = [
            str(ligne['rang']) if ligne['rang'] is not None else '-',
            ligne['etudiant'].last_name or '-',
            ligne['etudiant'].first_name or '-',
        ]
        for note in ligne['notes']:
            row.append(f"{note:.2f}" if note is not None else '-')
        row.append(f"{ligne['moyenne_20']:.2f}" if ligne['moyenne_20'] is not None else '-')
        row.append(f"{ligne['moyenne_10']:.2f}" if ligne['moyenne_10'] is not None else '-')
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(ENTETE_COULEUR_PDF)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f2ea")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ============================================================
# Resultats individuels (matieres validees / non validees)
# ============================================================

def _entete_resultats(etudiant, periode, etablissement):
    lignes_entete = []
    if etablissement.nom:
        entete_etab = etablissement.nom + (f" — Tél : {etablissement.telephone}" if etablissement.telephone else "")
        lignes_entete.append(entete_etab)
    lignes_entete.append(
        f"{etudiant.last_name or ''} {etudiant.first_name or ''}".strip() or etudiant.username
    )
    sous_titre = f"Période : {periode.nom}"
    if periode.annee_scolaire:
        sous_titre += f" — Année : {periode.annee_scolaire}"
    lignes_entete.append(sous_titre)
    return lignes_entete


def _lignes_resultats(liste_matieres):
    lignes = []
    for r in liste_matieres:
        lignes.append([
            str(r['matiere']) if r['matiere'] else 'Sans matière',
            ", ".join(r['enseignants']) or '-',
            float(r['total_points_obtenus']),
            float(r['total_points_possibles']),
            float(r['moyenne_20']) if r['moyenne_20'] is not None else '-',
            float(r['moyenne_10']) if r['moyenne_10'] is not None else '-',
        ])
    return lignes


def exporter_resultats_excel(etudiant, periode, resultats):
    etablissement = Etablissement.charger()
    entetes = ['Matière', 'Enseignant(s)', 'Points obtenus', 'Points possibles', 'Moyenne /20', 'Moyenne /10']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    ligne_courante = 1
    for texte in _entete_resultats(etudiant, periode, etablissement):
        ws.append([texte])
        ws.merge_cells(start_row=ligne_courante, start_column=1, end_row=ligne_courante, end_column=len(entetes))
        ws.cell(row=ligne_courante, column=1).font = Font(bold=True, size=12 if ligne_courante > 1 else 11)
        ligne_courante += 1
    ws.append([])
    ligne_courante += 1

    def ecrire_section(titre_section, liste_matieres):
        nonlocal ligne_courante
        ws.append([titre_section])
        ws.merge_cells(start_row=ligne_courante, start_column=1, end_row=ligne_courante, end_column=len(entetes))
        ws.cell(row=ligne_courante, column=1).font = Font(bold=True, size=12)
        ligne_courante += 1

        ws.append(entetes)
        for cell in ws[ligne_courante]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=ENTETE_COULEUR_XLSX)
        ligne_courante += 1

        for row in _lignes_resultats(liste_matieres):
            ws.append(row)
            ligne_courante += 1
        if not liste_matieres:
            ws.append(["Aucune matière"])
            ligne_courante += 1
        ws.append([])
        ligne_courante += 1

    ecrire_section("Matières validées", resultats['matieres_validees'])
    ecrire_section("Matières non validées", resultats['matieres_non_validees'])

    moy_g20 = resultats['moyenne_generale_20']
    moy_g10 = resultats['moyenne_generale_10']
    ws.append(["Moyenne générale", '', '', '', float(moy_g20) if moy_g20 is not None else '-', float(moy_g10) if moy_g10 is not None else '-'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    for idx in range(1, len(entetes) + 1):
        lettre = get_column_letter(idx)
        valeurs = [str(ws.cell(row=r, column=idx).value) for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=idx).value not in (None, '')]
        if valeurs:
            ws.column_dimensions[lettre].width = min(max(len(v) for v in valeurs) + 2, 35)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultats.xlsx"'
    wb.save(response)
    return response


def exporter_resultats_pdf(etudiant, periode, resultats):
    etablissement = Etablissement.charger()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="resultats.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    entete = _entete_resultats(etudiant, periode, etablissement)
    if etablissement.nom:
        elements.append(Paragraph(entete[0], styles['Heading3']))
        entete = entete[1:]
    elements.append(Paragraph(entete[0], styles['Title']))
    elements.append(Paragraph(entete[1], styles['Normal']))
    elements.append(Spacer(1, 12))

    entetes_table = ['Matière', 'Enseignant(s)', 'Pts obtenus', 'Pts possibles', 'Moy. /20', 'Moy. /10']

    def construire_table(liste_matieres):
        data = [entetes_table]
        for r in liste_matieres:
            data.append([
                str(r['matiere']) if r['matiere'] else 'Sans matière',
                ", ".join(r['enseignants']) or '-',
                f"{r['total_points_obtenus']:.2f}",
                f"{r['total_points_possibles']:.2f}",
                f"{r['moyenne_20']:.2f}" if r['moyenne_20'] is not None else '-',
                f"{r['moyenne_10']:.2f}" if r['moyenne_10'] is not None else '-',
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(ENTETE_COULEUR_PDF)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f2ea")]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        return table

    elements.append(Paragraph("Matières validées", styles['Heading2']))
    if resultats['matieres_validees']:
        elements.append(construire_table(resultats['matieres_validees']))
    else:
        elements.append(Paragraph("Aucune matière", styles['Normal']))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Matières non validées", styles['Heading2']))
    if resultats['matieres_non_validees']:
        elements.append(construire_table(resultats['matieres_non_validees']))
    else:
        elements.append(Paragraph("Aucune matière", styles['Normal']))
    elements.append(Spacer(1, 16))

    moy_g20 = resultats['moyenne_generale_20']
    moy_g10 = resultats['moyenne_generale_10']
    texte_moyenne = f"Moyenne générale : {moy_g20:.2f}/20 ({moy_g10:.2f}/10)" if moy_g20 is not None else "Moyenne générale : -"
    elements.append(Paragraph(texte_moyenne, styles['Heading3']))

    doc.build(elements)
    return response
